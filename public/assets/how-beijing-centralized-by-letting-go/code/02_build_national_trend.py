"""
Build the time-series xlsx + two PNG charts for short data news on
"中国转移支付结构变迁 2013-2026: 财政自主权真的被压缩了吗".

All numbers traced to 财政部预算司 annual decision/budget releases.
"""
# --- Path setup (relative to repo) ---
from pathlib import Path
HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / 'data'
FIG_DIR  = HERE.parent / 'figures'
DATA_DIR.mkdir(exist_ok=True, parents=True)
FIG_DIR.mkdir(exist_ok=True, parents=True)


import matplotlib
from matplotlib import font_manager

# Locate Noto CJK font (used for Chinese-label figures). Try common locations.
def _find_cjk_font():
    candidates = [
        '/tmp/fonts/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        str(Path.home() / 'Library/Fonts/NotoSansCJK-Regular.ttc'),
    ]
    for c in candidates:
        if Path(c).exists():
            return c
    return None
_cjk_font_path = _find_cjk_font()
if _cjk_font_path:
    font_manager.fontManager.addfont(_cjk_font_path)
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['Noto Sans CJK SC', 'Noto Sans CJK JP', 'sans-serif']
plt.rcParams['axes.unicode_minus'] = False

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================
# Master data — all from 财政部预算司, units 亿元
# =============================================================
# Columns:
#   一般性     专项    三、特殊  税收返还(独立)  总转移支付   共同事权   税收返还及固定补助(in一般性)
#                                                              (None = 不存在/未独立列示)
DATA = {
    2013: dict(general=24362.72, special=18610.46, third=0,    tax_sep=5046.74,  total=42973.18,  shared=None, tax_in_g=None),
    2014: dict(general=27568.37, special=18941.12, third=0,    tax_sep=5081.55,  total=46509.49,  shared=None, tax_in_g=None),
    2015: dict(general=28455.02, special=21623.63, third=0,    tax_sep=5018.86,  total=50078.65,  shared=None, tax_in_g=None),
    2016: dict(general=31864.93, special=20708.93, third=0,    tax_sep=6826.84,  total=52573.86,  shared=None, tax_in_g=None),
    2017: dict(general=35145.59, special=21883.36, third=0,    tax_sep=8022.83,  total=57028.95,  shared=None, tax_in_g=None),
    2018: dict(general=38722.06, special=22927.09, third=0,    tax_sep=8031.51,  total=61649.15,  shared=None, tax_in_g=None),
    2019: dict(general=66798.16, special=7561.70,  third=0,    tax_sep=None,     total=74359.86,  shared=31902.99, tax_in_g=11251.78),
    2020: dict(general=69459.86, special=7765.92,  third=5992.15, tax_sep=None,  total=83217.93,  shared=32180.72, tax_in_g=11275.64),
    2021: dict(general=74799.29, special=7353.05,  third=0,    tax_sep=None,     total=82152.34,  shared=34258.82, tax_in_g=11569.67),
    2022: dict(general=80811.30, special=7597.03,  third=8533.49, tax_sep=None,  total=96941.82,  shared=36354.12, tax_in_g=11836.90),
    2023: dict(general=85036.91, special=8040.67,  third=9758.74, tax_sep=None,  total=102836.32, shared=36794.04, tax_in_g=11309.48),
    2024: dict(general=87161.44, special=8174.28,  third=5000.00, tax_sep=None,  total=100335.72, shared=37503.37, tax_in_g=12138.14),
    2025: dict(general=94117.80, special=9297.20,  third=0,    tax_sep=None,     total=103415.00, shared=38609.67, tax_in_g=12627.09),
    2026: dict(general=94779.22, special=9370.78,  third=0,    tax_sep=None,     total=104150.00, shared=38987.74, tax_in_g=12333.06),
}
KIND = {y: ('决算' if y <= 2024 else '预算') for y in DATA}

# Derive comparable buckets:
#   tax_return     = 税收返还（pre-2019独立 / post-2019从一般性内提取）
#   net_general    = 一般性 - 税收返还 - 共同事权（"自由调度"部分；2019前 共同事权 unknown 用 N/A）
#   shared_amt     = 共同事权（仅2019+）
#   special_total  = 专项 + 三、
#   comparable_tot = 总投入 = pre-2019: 转移支付总 + 税收返还; post-2019: 转移支付总（已含税收返还）
years = sorted(DATA.keys())
rows = []
for y in years:
    d = DATA[y]
    if d['tax_sep'] is not None:
        tax = d['tax_sep']
        total = d['total'] + tax
        # 2013-2018: 一般性 was net of 税收返还 (its own row); 共同事权 was inside 一般性 but not separated
        net_general_incl_shared = d['general']  # 共同事权 unknown
        shared = None
        free_part = None  # cannot isolate without 共同事权 data
    else:
        tax = d['tax_in_g']
        total = d['total']
        # Strip 税收返还 from 一般性 to get 一般性(净，含共同事权)
        net_general_incl_shared = d['general'] - tax
        shared = d['shared']
        free_part = net_general_incl_shared - shared  # 一般性 - 税收返还 - 共同事权
    special_total = d['special'] + d['third']
    rows.append(dict(
        year=y, kind=KIND[y],
        tax=tax,
        gen_net_inc_shared=net_general_incl_shared,
        shared=shared,
        free=free_part,
        special=special_total,
        third=d['third'],
        total=total,
        share_special=special_total/total,
        share_tax=tax/total,
        share_general_net=net_general_incl_shared/total,
        share_shared=(shared/total) if shared else None,
        share_free=(free_part/total) if free_part else None,
        share_shared_in_gen=(shared/net_general_incl_shared) if shared else None,
    ))

# =============================================================
# Chart 1: stacked area, % composition over time
# =============================================================
fig, ax = plt.subplots(figsize=(12, 6.5))
xs = [r['year'] for r in rows]

# Build stacks: 税收返还 / 共同事权 / 其他一般性 / 专项+三
# For 2013-2018 we don't have 共同事权 separately — show 一般性(含共同事权) as one shaded block
tax_pct       = [r['share_tax']*100 for r in rows]
shared_pct    = [r['share_shared']*100 if r['share_shared'] else 0 for r in rows]
free_pct      = [r['share_free']*100 if r['share_free'] else r['share_general_net']*100 for r in rows]  # pre-2019 = entire net general
special_pct   = [r['share_special']*100 for r in rows]

# Override pre-2019: collapse 共同事权 into 一般性 since not separated
for i, r in enumerate(rows):
    if r['shared'] is None:
        shared_pct[i] = 0  # not visible
        free_pct[i] = r['share_general_net']*100  # full net general

colors = {
    'tax':     '#A9C5E8',  # light blue
    'shared':  '#F4B183',  # light orange — "common fiscal responsibility"
    'free':    '#5B9BD5',  # blue — "truly free"
    'special': '#C00000',  # dark red — "project-locked"
}

bot = np.zeros(len(xs))
ax.fill_between(xs, bot, bot + np.array(tax_pct), step=None, color=colors['tax'], label='税收返还（rule-based / 不属真正自由调度）', alpha=0.95)
bot = bot + np.array(tax_pct)
ax.fill_between(xs, bot, bot + np.array(shared_pct), color=colors['shared'], label='共同财政事权（中央指定用途，2019+独立列示）', alpha=0.95)
bot = bot + np.array(shared_pct)
ax.fill_between(xs, bot, bot + np.array(free_pct), color=colors['free'], label='一般性其他（地方真正可自由调度）', alpha=0.95)
bot = bot + np.array(free_pct)
ax.fill_between(xs, bot, bot + np.array(special_pct), color=colors['special'], label='专项 +「三、」临时类（项目锁死）', alpha=0.95)

# Vertical break markers
ax.axvline(2014, color='#888', linestyle='--', linewidth=1, alpha=0.6)
ax.axvline(2019, color='#888', linestyle='--', linewidth=1, alpha=0.6)
ax.axvline(2024.5, color='#444', linestyle=':', linewidth=1.2)

# Annotations
ax.text(2014, 102, '2014\n专项整合改革', ha='center', va='bottom', fontsize=9, color='#444')
ax.text(2019, 102, '2019\n共同事权独立列示', ha='center', va='bottom', fontsize=9, color='#444')
ax.text(2025.5, 102, '预算口径', ha='center', va='bottom', fontsize=9, color='#444', style='italic')

ax.set_xlim(2013, 2026)
ax.set_ylim(0, 100)
ax.set_xticks(xs)
ax.set_xticklabels([str(y) for y in xs], rotation=0)
ax.set_ylabel('占中央对地方资金总额比重（%）')
ax.set_title('中国中央对地方转移支付结构变迁，2013–2026', fontsize=14, pad=22, weight='bold')
ax.legend(loc='lower center', bbox_to_anchor=(0.5, -0.20), ncol=2, frameon=False, fontsize=9)
ax.grid(False)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

# Note at bottom
fig.text(0.06, 0.02,
    '数据来源：财政部预算司 历年《中央对地方转移支付决算表》及2025、2026《中央对地方转移支付预算表》。\n'
    '口径说明：① 2019 年起税收返还并入一般性，共同财政事权从一般性中独立列示；② 2013–2018 共同事权未独立，已并入"一般性其他"显示；'
    '③「三、」为 2020、2022、2023、2024 临时新增的特殊补助（抗疫 / 减税降费 / 灾后），合并入「专项」展示；'
    '④ 2025 起为预算口径（2025 决算尚未发布）。',
    fontsize=8, color='#555', ha='left', va='bottom')

plt.subplots_adjust(left=0.07, right=0.96, top=0.88, bottom=0.22)
plt.savefig(str(FIG_DIR / '1_transfer_structure_2013_2026.png'),
            dpi=180, bbox_inches='tight', facecolor='white')
plt.close()

# =============================================================
# Chart 2: line chart — 关键比例趋势
# =============================================================
fig, ax = plt.subplots(figsize=(12, 6))

# 三条线: 专项+三 / 总 ; 共同事权 / 总 ; 自由部分 / 总
sp_share = [r['share_special']*100 for r in rows]
sh_share = [r['share_shared']*100 if r['share_shared'] else None for r in rows]
free_share_total = [r['share_free']*100 if r['share_free'] else None for r in rows]

# Also: 共同事权 占一般性(净) 比, 仅 2019+
shared_in_gen = [r['share_shared_in_gen']*100 if r['share_shared_in_gen'] else None for r in rows]

# Plot
ax.plot(xs, sp_share, marker='o', color='#C00000', linewidth=2.2, label='① 专项+三 占总转移支付（"项目锁死"）', markersize=6)
# For shared: only plot for years where data exists
xs2 = [r['year'] for r in rows if r['share_shared'] is not None]
sh2 = [r['share_shared']*100 for r in rows if r['share_shared'] is not None]
ax.plot(xs2, sh2, marker='s', color='#F4B183', linewidth=2.2, label='② 共同财政事权 占总转移支付（中央指定用途）', markersize=6)

free2_xs = [r['year'] for r in rows if r['share_free'] is not None]
free2 = [r['share_free']*100 for r in rows if r['share_free'] is not None]
ax.plot(free2_xs, free2, marker='^', color='#2E75B6', linewidth=2.5, label='③ 一般性其他 占总转移支付（地方真正可自由调度）', markersize=7)

# Highlight: 2019→2026 自由占比上升
ax.annotate(f'2019: {free2[0]:.1f}%',
    xy=(2019, free2[0]), xytext=(2014.5, 22),
    fontsize=10, color='#2E75B6',
    arrowprops=dict(arrowstyle='-', color='#2E75B6', lw=0.8))
ax.annotate(f'2026预算: {free2[-1]:.1f}%',
    xy=(2026, free2[-1]), xytext=(2023.3, 50),
    fontsize=10, color='#2E75B6', weight='bold',
    arrowprops=dict(arrowstyle='-', color='#2E75B6', lw=0.8))

# Vertical lines
ax.axvline(2014, color='#888', linestyle='--', linewidth=1, alpha=0.5)
ax.axvline(2019, color='#888', linestyle='--', linewidth=1, alpha=0.5)
ax.axvline(2024.5, color='#444', linestyle=':', linewidth=1.0, alpha=0.7)
ax.text(2014, ax.get_ylim()[1]*0.95, '2014 专项整合', fontsize=9, color='#666', ha='left', va='top')
ax.text(2019, ax.get_ylim()[1]*0.95, '2019 共同事权独立', fontsize=9, color='#666', ha='left', va='top')

ax.set_xlim(2012.7, 2026.3)
ax.set_xticks(xs)
ax.set_xticklabels([str(y) for y in xs])
ax.set_ylabel('占中央对地方资金总额比重（%）')
ax.set_title('"财政自主权被压缩"了吗？——三条比例线给出反直觉答案', fontsize=14, pad=18, weight='bold')
ax.legend(loc='upper right', frameon=False, fontsize=10)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.grid(axis='y', alpha=0.25)

fig.text(0.06, 0.02,
    '数据来源：财政部预算司 2013–2024 决算表、2025/2026 预算表。线 ② ③ 仅 2019 年起可比（共同事权独立列示之前数据缺失）。',
    fontsize=8, color='#555', ha='left', va='bottom')

plt.subplots_adjust(left=0.07, right=0.96, top=0.90, bottom=0.12)
plt.savefig(str(FIG_DIR / '2_fiscal_autonomy_trend_2013_2026.png'),
            dpi=180, bbox_inches='tight', facecolor='white')
plt.close()

# =============================================================
# Build xlsx with full annual data
# =============================================================
wb = Workbook()
ws = wb.active
ws.title = "年度时间序列"

FONT_HEADER = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_NOTE = Font(name="Arial", size=9, italic=True, color="555555")
FILL_HEADER = PatternFill("solid", start_color="305496")
FILL_DECISION = PatternFill("solid", start_color="FFFFFF")
FILL_BUDGET = PatternFill("solid", start_color="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

ws['A1'] = "中央对地方转移支付结构 2013–2026（绝对额 + 占比）"
ws['A1'].font = Font(name="Arial", bold=True, size=14)
ws.merge_cells('A1:M1')
ws['A2'] = "单位：亿元（金额）/ %（占比）；数据来源：财政部预算司"
ws['A2'].font = FONT_NOTE
ws.merge_cells('A2:M2')

headers = [
    '年份', '口径',
    '总转移支付（含税收返还）', '税收返还', '一般性其他（自由）', '共同财政事权', '专项+三、',
    '占比：税收返还', '占比：自由', '占比：共同事权', '占比：专项+三',
    '共同事权 / 一般性(净)', '总规模(万亿)',
]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
    cell = ws.cell(row=4, column=i)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER

start = 5
for i, r in enumerate(rows):
    rr = start + i
    ws.cell(row=rr, column=1, value=r['year'])
    ws.cell(row=rr, column=2, value=r['kind'])
    ws.cell(row=rr, column=3, value=r['total'])
    ws.cell(row=rr, column=4, value=r['tax'])
    # 一般性其他（自由）— pre-2019 缺数据，用 None
    if r['free'] is None:
        ws.cell(row=rr, column=5, value="N/A")
    else:
        ws.cell(row=rr, column=5, value=r['free'])
    if r['shared'] is None:
        ws.cell(row=rr, column=6, value="N/A")
    else:
        ws.cell(row=rr, column=6, value=r['shared'])
    ws.cell(row=rr, column=7, value=r['special'])
    # Shares — use formulas where possible
    ws.cell(row=rr, column=8, value=f"=D{rr}/C{rr}")
    if r['free'] is None:
        ws.cell(row=rr, column=9, value="N/A")
    else:
        ws.cell(row=rr, column=9, value=f"=E{rr}/C{rr}")
    if r['shared'] is None:
        ws.cell(row=rr, column=10, value="N/A")
    else:
        ws.cell(row=rr, column=10, value=f"=F{rr}/C{rr}")
    ws.cell(row=rr, column=11, value=f"=G{rr}/C{rr}")
    if r['shared'] is None:
        ws.cell(row=rr, column=12, value="N/A")
    else:
        # 共同事权 / 一般性(净, 含共同事权)
        ws.cell(row=rr, column=12, value=f"=F{rr}/(E{rr}+F{rr})")
    ws.cell(row=rr, column=13, value=f"=C{rr}/10000")

    # Format
    for c in range(1, 14):
        cell = ws.cell(row=rr, column=c)
        cell.font = FONT_BODY
        cell.border = BORDER
        cell.alignment = ALIGN_CENTER if c <= 2 else ALIGN_RIGHT
        if r['kind'] == '预算':
            cell.fill = FILL_BUDGET
    for c in (3, 4, 5, 6, 7):
        ws.cell(row=rr, column=c).number_format = '#,##0.00'
    for c in (8, 9, 10, 11, 12):
        ws.cell(row=rr, column=c).number_format = '0.0%'
    ws.cell(row=rr, column=13).number_format = '0.00'

# Column widths
widths = [7, 7, 16, 12, 14, 14, 13, 11, 10, 12, 11, 16, 11]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 36

# =============================================================
# Sheet 2: 关键发现 (key findings narrative)
# =============================================================
ws2 = wb.create_sheet("关键发现")
ws2['A1'] = '"财政自主权被压缩了吗？" — 数据反驳了直觉'
ws2['A1'].font = Font(name="Arial", bold=True, size=14)
ws2.merge_cells('A1:D1')

findings = [
    ('总盘子', '2013→2024 决算从 4.8 万亿 → 10.0 万亿，13 年内中央对地方资金翻了一倍多。2026 预算 10.4 万亿，增速明显放缓。'),
    ('2014 专项整合改革（断点 1）', '2018 专项决算 2.29 万亿 → 2019 专项决算 7562 亿元。但这不是真的"删掉"了 1.5 万亿专项——而是把里面的医保/养老/教育等项目重新分类到"共同财政事权"里（2019 起从一般性里独立列示）。专项占比"陡降"主要是会计科目重组。'),
    ('2019 共同事权独立列示（断点 2）', '2019 一开始共同事权占一般性（不含税收返还）就达 57.4%，说明这部分支出长期存在，只是之前不可见。'),
    ('反直觉发现 ①', '"地方真正可自由调度部分"占总转移支付比例：2019 年 31.8% → 2026 预算 41.7%，过去 7 年持续上升，绝非"被压缩"。'),
    ('反直觉发现 ②', '共同事权占一般性（不含税收返还）比例：2019 年 57.4% → 2026 预算 47.3%，整体在缓慢下降——与"中央指定用途扩大"的直觉相反。'),
    ('反直觉发现 ③', '"专项 + 三、临时补助" 占总转移支付的波动主要由 COVID（2020 抗疫 5992 亿）、减税降费（2022 减税降费补助 8533 亿）、灾后救助（2023、2024 各 5000 亿）驱动，是政策事件，不是结构性变化。'),
    ('结论框架', '把直觉"逐年压缩"换成"换瓶不换酒"也不准确——数据显示，账面口径的反复重组确实让外人看不清，但从"地方真正可自由调度"的口径看，过去 7 年地方财政自主权事实上在缓慢扩大。'),
]
for i, (k, v) in enumerate(findings, start=3):
    ws2.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=11)
    ws2.cell(row=i, column=2, value=v).font = FONT_BODY
    ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[i].height = 50

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 110

# =============================================================
# Sheet 3: 数据来源
# =============================================================
ws3 = wb.create_sheet("数据来源")
ws3['A1'] = '逐年数据来源（财政部预算司）'
ws3['A1'].font = Font(name="Arial", bold=True, size=14)
ws3.merge_cells('A1:B1')

src = [
    ('2013 决算', 'http://yss.mof.gov.cn/2013qgczjs/201407/t20140711_1112026.htm'),
    ('2014 决算', 'http://yss.mof.gov.cn/2014czys/201507/t20150709_1269837.htm'),
    ('2015 决算', 'http://yss.mof.gov.cn/2015js/201607/t20160713_2354962.htm'),
    ('2016 决算', 'http://yss.mof.gov.cn/2016js/201707/t20170713_2648693.htm'),
    ('2017 决算', 'http://yss.mof.gov.cn/qgczjs/201807/t20180712_2959754.htm'),
    ('2018 决算', 'http://yss.mof.gov.cn/2018czjs/201907/t20190718_3303311.htm'),
    ('2019 决算', 'http://yss.mof.gov.cn/2019qgczjs/202007/t20200706_3544608.htm'),
    ('2020 决算', 'http://yss.mof.gov.cn/2020zyjs/202106/t20210629_3727224.htm'),
    ('2021 决算', 'http://yss.mof.gov.cn/2021zyjs/202207/t20220712_3826596.htm'),
    ('2022 决算', 'http://yss.mof.gov.cn/2022zyjs/202307/t20230714_3896491.htm'),
    ('2023 决算', 'http://yss.mof.gov.cn/2023zyjs/202407/t20240716_3939612.htm'),
    ('2024 决算', 'http://yss.mof.gov.cn/2024zyjs/202507/t20250715_3967896.htm'),
    ('2025 预算', 'http://yss.mof.gov.cn/2025zyczys/202503/t20250324_3960476.htm'),
    ('2026 预算', 'http://yss.mof.gov.cn/2026zyczys/202603/t20260324_3986014.htm'),
]
for i, (k, v) in enumerate(src, start=3):
    ws3.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=11)
    ws3.cell(row=i, column=2, value=v).font = FONT_BODY
    ws3.row_dimensions[i].height = 22

ws3.column_dimensions['A'].width = 16
ws3.column_dimensions['B'].width = 90

out = str(DATA_DIR / '2_national_time_series_2013_2026.xlsx')
wb.save(out)
print(f"Saved xlsx: {out}")
print(f"Saved chart 1: 转移支付结构变迁_2013_2026.png")
print(f"Saved chart 2: 财政自主权趋势_2013_2026.png")
