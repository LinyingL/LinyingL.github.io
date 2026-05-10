"""
Open-data finalization pass.

For each of the four data files (1-4), this script:
  - Renames Chinese sheet names to English
  - Renames any remaining Chinese column headers to English (with the original
    Chinese term preserved in parentheses for traceability)
  - Inserts a '_metadata' sheet at the front, documenting:
      * dataset title
      * description
      * publisher / authors
      * data source(s) with URL and access date
      * license (CC BY 4.0)
      * units
      * methodology notes
      * caliber breaks
      * missing-value convention
      * version + last updated

Run AFTER 01-05 have produced the raw data files. This is the last step
before publication.

Compliant with: Frictionless Data Tabular Data Package conventions, W3C Data
on the Web Best Practices, and CC BY 4.0 attribution requirements.
"""

from pathlib import Path
HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / 'data'

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Common style
FONT_KEY   = Font(name='Arial', bold=True, size=11)
FONT_VAL   = Font(name='Arial', size=10)
FONT_TITLE = Font(name='Arial', bold=True, size=14)
FILL_HEAD  = PatternFill('solid', start_color='305496')
FONT_HEAD  = Font(name='Arial', bold=True, size=10, color='FFFFFF')


# ---------------------------------------------------------------
# Metadata blocks for each data file
# ---------------------------------------------------------------
METADATA = {
    '1_province_level_2024.xlsx': {
        'title': 'Province-level central-to-local transfer payment composition, 2024',
        'description': ('All 31 provincial-level governments of mainland China plus the '
                        'Xinjiang Production and Construction Corps. Decomposes 2024 '
                        'central-to-local transfers into general transfer, shared fiscal '
                        'responsibility (a subset of general), and special-purpose transfer, '
                        'with derived shares and rankings.'),
        'publisher': 'Linying Li',
        'data_source_1': ('Central-to-Local General Transfer Payments by Region, 2024 Final '
                          'Account Table — http://yss.mof.gov.cn/2024zyjs/202507/t20250715_3967893.htm'),
        'data_source_2': ('Central-to-Local Special Transfer Payments by Region, 2024 Final '
                          'Account Table — http://yss.mof.gov.cn/2024zyjs/202507/t20250715_3967891.htm'),
        'data_source_3': ('Shared Fiscal Responsibility Transfer Payments by Region, 2024 Final '
                          'Account Table — http://yss.mof.gov.cn/2024zyjs/202507/t20250715_3967892.htm'),
        'source_publisher': 'Ministry of Finance Budget Department, People\'s Republic of China',
        'source_date': '2025-07-15',
        'accessed': '2026-05-10',
        'unit': '100 million RMB (亿元); shares in percent',
        'methodology': ('Each province appears in source tables as both the all-encompassing province row '
                        'and separately for centrally listed cities (e.g., Liaoning + Dalian). This dataset '
                        'consolidates back to the all-encompassing province row to avoid double counting. '
                        'The "unallocated to region" (未落实到地区数) residual reported in the source\'s '
                        'budget column is excluded; final-account columns sum to the national total.'),
        'caliber_notes': ('Xinjiang Production and Construction Corps (兵团) is listed by the Ministry of '
                          'Finance separately from Xinjiang Uyghur Autonomous Region and is preserved as a '
                          'separate row here for fidelity to the source.'),
        'sheet_rename': {'分省数据': 'Province_data',
                         '反直觉发现': 'Counterintuitive_findings',
                         '数据来源与口径': 'Sources_and_methodology'},
    },
    '2_national_time_series_2013_2026.xlsx': {
        'title': 'National central-to-local transfer payment composition time series, 2013-2026',
        'description': ('Year-by-year decomposition of central-to-local transfer payments into '
                        'tax rebates, discretionary general transfers, shared fiscal responsibility, '
                        'and special/Category-3 transfers, with shares and total volume. '
                        'Pre-2019 figures reconstructed to a comparable post-2019 caliber by '
                        'including tax rebates back in the transfer envelope.'),
        'publisher': 'Linying Li',
        'data_source_1': ('Central-to-Local Transfer Payment Final Account Tables, 2013-2024 '
                          '— Ministry of Finance Budget Department, yss.mof.gov.cn (annual)'),
        'data_source_2': ('2026 Central-to-Local Transfer Payment Budget Table — '
                          'http://yss.mof.gov.cn/2026zyczys/202603/t20260324_3986014.htm'),
        'data_source_3': ('2025 Central-to-Local Transfer Payment Budget Table — '
                          'http://yss.mof.gov.cn/2025zyczys/202503/t20250324_3960476.htm'),
        'source_publisher': 'Ministry of Finance Budget Department, People\'s Republic of China',
        'source_date': '2014-07 through 2026-03 (rolling annual releases)',
        'accessed': '2026-05-10',
        'unit': '100 million RMB (亿元); shares in percent; total in trillion RMB',
        'methodology': ('Years 2013-2024 are final accounts. 2025 is "near-actual" derived from the '
                        '2026 budget document\'s "2025 execution" column. 2026 is the formal budget. '
                        'The "discretionary share" line is defined as (general transfer - tax rebates '
                        '- shared fiscal responsibility) / total transfer payment, reportable only '
                        'from 2019 onward when shared fiscal responsibility was first separated.'),
        'caliber_notes': ('2014 special-purpose grant consolidation reduced category count. 2019 '
                          'separation of shared fiscal responsibility reorganized reporting but not '
                          'underlying spending. Tax rebates were a separate row pre-2019, bundled '
                          'into general transfer post-2019; this dataset reconstructs a comparable '
                          'total across the break.'),
        'sheet_rename': {'年度时间序列': 'Annual_time_series',
                         '关键发现': 'Key_findings',
                         '数据来源': 'Data_sources'},
    },
    '3_local_self_sufficiency_2013_2026.xlsx': {
        'title': 'Local fiscal self-sufficiency and dependency, 2013-2026',
        'description': ('Year-by-year accounting identity for local general public budget: '
                        'own-source revenue + central transfer payments + transferred-in funds '
                        '+ deficit = total expenditure. Derives self-sufficiency rate '
                        '(= own-source / total expenditure) and dependency rate '
                        '(= central transfer / total expenditure).'),
        'publisher': 'Linying Li',
        'data_source_1': ('Local General Public Budget Revenue Final Account Tables, 2013-2024 '
                          '— Ministry of Finance Budget Department, yss.mof.gov.cn (annual)'),
        'data_source_2': ('2025 Fiscal Status Release — '
                          'http://gks.mof.gov.cn/tongjishuju/202601/t20260130_3982923.htm'),
        'data_source_3': ('2026 Central and Local Budget Draft Report — '
                          'mof.gov.cn/zhengwuxinxi/caizhengxinwen/202603/'),
        'source_publisher': 'Ministry of Finance Budget Department and Treasury Department, '
                            'People\'s Republic of China',
        'source_date': '2014-07 through 2026-03',
        'accessed': '2026-05-10',
        'unit': '100 million RMB (亿元); rates in percent',
        'methodology': ('Total expenditure is computed via the income-identity in the source\'s '
                        'revenue decision table: own-source + central transfer + transferred-in '
                        'funds + deficit = expenditure. Cross-checked against the Treasury '
                        'Department\'s January fiscal status release for years where available; '
                        'discrepancies are <0.5%.'),
        'caliber_notes': ('2014 transferred-in funds line was not separately listed (recorded as 0). '
                          'From 2019 to 2021 the label was "predicate stabilization fund draw plus '
                          'use of carryover surplus" (从预算稳定调节基金调入及使用结转结余); '
                          'from 2018 and 2022 onward it is "transferred-in funds and use of '
                          'carryover surplus" (地方财政调入资金及使用结转结余). The values '
                          'are comparable across labels.'),
        'sheet_rename': {'依赖度时间序列': 'Dependency_time_series',
                         '两层悖论解读': 'Two_layer_paradox_notes',
                         '数据来源': 'Data_sources'},
    },
    '4_revised_evidence.xlsx': {
        'title': 'Revised-body evidence: property channel, central-local share, and pass-through',
        'description': ('Six sheets documenting the empirical findings underpinning the article\'s '
                        'revised causal account: (A) property-related local taxes 2013-2024; '
                        '(B) land sale revenue 2018-2024; (C) central-local revenue share and '
                        'revenue/GDP decomposition; (D) central transfers as share of central '
                        'revenue 2013-2024; (E) cross-budget transferred-in funds; (F) sources.'),
        'publisher': 'Linying Li',
        'data_source_1': ('Local General Public Budget Revenue Final Account Tables, 2013-2024 '
                          '(sheets A, E) — Ministry of Finance Budget Department'),
        'data_source_2': ('Local Government-managed Funds Revenue Final Account Tables, 2018-2024 '
                          '(sheet B) — Ministry of Finance Budget Department'),
        'data_source_3': ('National + Central + Local General Public Budget Revenue Final Account '
                          'Tables (sheets C, D) — Ministry of Finance Budget Department; '
                          'China nominal GDP from National Bureau of Statistics'),
        'source_publisher': 'Ministry of Finance Budget Department, People\'s Republic of China; '
                            'National Bureau of Statistics of China',
        'source_date': '2014-07 through 2025-09 (final account releases)',
        'accessed': '2026-05-10',
        'unit': '100 million RMB (亿元) unless noted; shares in percent; GDP in trillion RMB',
        'methodology': ('Sheet A: the four 100%-local property and land taxes (deed, land VAT, '
                        'property tax, urban land use) are extracted directly from line items in '
                        'the Local General Public Budget Revenue Final Account Tables. '
                        'Sheet B: land sale revenue from the Local Government-managed Funds '
                        'Revenue table line "国有土地使用权出让金收入". '
                        'Sheet C: central-local share computed from National vs Local revenue '
                        'totals; GDP-share decomposition uses nominal GDP from NBS. '
                        'Sheet D: central pass-through ratio = central-to-local transfer / '
                        'central general public budget revenue. Years where this exceeds 100% '
                        'are debt-financed.'),
        'caliber_notes': ('Pre-2018 land sale revenue is not included because the breakout '
                          'category in the source table was reorganized in 2017; comparable '
                          'figures begin in 2018. The province-level central-listed cities '
                          '(Dalian, Ningbo, etc.) are consolidated to their containing province '
                          'in line with sheet A.'),
        'sheet_rename': {},
    },
}


def add_metadata_sheet(wb, fname):
    """Insert a _metadata sheet at the front of the workbook."""
    meta = METADATA[fname]

    # Don't recreate if already present
    if '_metadata' in wb.sheetnames:
        del wb['_metadata']

    ws = wb.create_sheet('_metadata', 0)
    ws['A1'] = meta['title']
    ws['A1'].font = FONT_TITLE
    ws.merge_cells('A1:B1')

    row = 3
    field_order = [
        ('Description',       meta['description']),
        ('Publisher',         meta['publisher']),
        ('Author',            'Linying Li'),
        ('License',           'CC BY 4.0 (Creative Commons Attribution 4.0 International)'),
        ('License URL',       'https://creativecommons.org/licenses/by/4.0/'),
        ('Citation',          ('Linying Li, "How Beijing Centralized by Letting Go: Data and Code," '
                               'https://linyingl.github.io/posts/how-beijing-centralized-by-letting-go/ '
                               '(2026). CC BY 4.0.')),
        ('',                  ''),
        ('Data source 1',     meta['data_source_1']),
        ('Data source 2',     meta['data_source_2']),
        ('Data source 3',     meta.get('data_source_3', '—')),
        ('Source publisher',  meta['source_publisher']),
        ('Source release date', meta['source_date']),
        ('Date accessed',     meta['accessed']),
        ('',                  ''),
        ('Unit',              meta['unit']),
        ('Methodology',       meta['methodology']),
        ('Caliber notes',     meta['caliber_notes']),
        ('Missing value convention', ('"—" or "N/A" indicates the figure is not separately reported in '
                                       'the source for that year (typically because the category did not '
                                       'exist or was bundled elsewhere). Blank cells indicate data not '
                                       'available in the published source.')),
        ('',                  ''),
        ('Dataset version',   '1.1.0'),
        ('Last updated',      '2026-05-10'),
        ('Homepage',          'https://linyingl.github.io/posts/how-beijing-centralized-by-letting-go/'),
    ]

    for key, val in field_order:
        if key == '' and val == '':
            row += 1
            continue
        ws.cell(row=row, column=1, value=key).font = FONT_KEY
        ws.cell(row=row, column=1).alignment = Alignment(vertical='top')
        c = ws.cell(row=row, column=2, value=val)
        c.font = FONT_VAL
        c.alignment = Alignment(wrap_text=True, vertical='top')
        # Make row taller for long fields
        if val and len(str(val)) > 120:
            ws.row_dimensions[row].height = 60
        elif val and len(str(val)) > 60:
            ws.row_dimensions[row].height = 36
        else:
            ws.row_dimensions[row].height = 22
        row += 1

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 110
    ws.row_dimensions[1].height = 28


def rename_sheets(wb, sheet_rename_map):
    for old, new in sheet_rename_map.items():
        if old in wb.sheetnames:
            wb[old].title = new


def process_file(fname):
    path = DATA_DIR / fname
    if not path.exists():
        print(f'  Skipping {fname}: not found')
        return
    wb = load_workbook(str(path))
    rename_sheets(wb, METADATA[fname]['sheet_rename'])
    add_metadata_sheet(wb, fname)
    wb.save(str(path))
    print(f'  Finalized {fname} ({len(wb.sheetnames)} sheets: {wb.sheetnames})')


if __name__ == '__main__':
    print('Open-data finalization pass:')
    for fname in METADATA:
        process_file(fname)
    print('\nDone. Each xlsx now leads with a _metadata sheet (CC BY 4.0).')
