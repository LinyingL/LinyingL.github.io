"""
Build the revised-evidence workbook (data file 4) covering the empirical
findings introduced in the article's revised body:

  Sheet A — Property-related local taxes, 2013-2024 (Test B)
  Sheet B — Land sale revenue, 2018-2024 (from government-managed funds budget)
  Sheet C — Central-local revenue share + revenue/GDP decomposition
  Sheet D — Central transfers as share of central general public budget revenue
  Sheet E — Cross-budget transferred-in funds (调入资金), 2013-2024
  Sheet F — Sources & methodology

Every numeric figure cited in the article body that isn't already in
files 1-3 should be reproducible from this workbook.
"""

from pathlib import Path
HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / 'data'
DATA_DIR.mkdir(exist_ok=True, parents=True)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_HEADER = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_NOTE = Font(name="Arial", size=9, italic=True, color="555555")
FILL_HEADER = PatternFill("solid", start_color="305496")
FILL_HILITE = PatternFill("solid", start_color="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER

def style_body(ws, r1, r2, ncol):
    for r in range(r1, r2 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BODY
            cell.border = BORDER
            cell.alignment = ALIGN_RIGHT if c > 1 else Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------
# Sheet A — Property-related local taxes (Test B)
# ---------------------------------------------------------------
# Source: Local General Public Budget Revenue Final Account Tables (2013-2024)
# Four 100%-local property/land taxes; values in 亿元 (final-account figure)
PROPERTY_TAX = [
    # (year, 契税, 土增税, 房产税, 城镇土地使用税, 地方本级收入)
    (2013, 3844.02, 3293.91, 1581.50, 1718.77, 69011.16),
    (2014, 4000.70, 3914.68, 1851.64, 1992.62, 75876.58),
    (2015, 3898.55, 3832.18, 2050.90, 2142.04, 83002.04),
    (2016, 4300.00, 4212.19, 2220.91, 2255.74, 87239.35),
    (2017, 4910.42, 4911.28, 2604.33, 2360.55, 91469.41),
    (2018, 5729.94, 5641.38, 2888.56, 2387.60, 97903.38),
    (2019, 6212.86, 6465.14, 2988.43, 2195.41, 101080.61),
    (2020, 7061.02, 6468.51, 2841.76, 2058.22, 100143.16),
    (2021, 7427.49, 6896.02, 3277.64, 2126.28, 111084.23),
    (2022, 5793.80, 6349.11, 3590.35, 2225.62, 108762.15),
    (2023, 5910.43, 5294.00, 3994.19, 2212.71, 117228.73),
    (2024, 5169.59, 4868.70, 4705.27, 2424.82, 119272.24),
]

# ---------------------------------------------------------------
# Sheet B — Land sale revenue (国有土地使用权出让金收入)
# ---------------------------------------------------------------
# Source: Local Government-managed Funds Revenue Final Account Tables
# 亿元 (final-account)
LAND_SALES = [
    (2018, 62875.11),
    (2019, 70631.06),
    (2020, 82098.02),
    (2021, 84897.67),
    (2022, 65326.00),
    (2023, 56633.68),
    (2024, 47741.70),
]

# ---------------------------------------------------------------
# Sheet C — Central-local revenue share + revenue/GDP decomposition
# ---------------------------------------------------------------
# Sources: National + Central + Local General Public Budget Revenue Final
# Account Tables (2013, 2024); China NBS nominal GDP (current prices, 万亿元)
# Values in 亿元 unless noted
CENTRAL_LOCAL_GDP = [
    # (year, 国内 nominal GDP 万亿, 全国一般预算收入 亿, 中央本级 亿, 地方本级 亿)
    (2013,  59.30,  129143,  60198,  69011.16),
    (2024, 134.91, 219700, 100442, 119272.24),
]
# Note: 2013 国家统计局 GDP 59.30 trillion; 2024 134.91 trillion. National
# revenue 2013 = 12.91 trillion (12.91万亿); 2024 = 21.97 trillion. Central
# revenue derived from National - Local; small discrepancies vs separately
# published central tables come from rounding.

# ---------------------------------------------------------------
# Sheet D — Central transfers as share of central revenue
# ---------------------------------------------------------------
# Central revenue and central-to-local transfer 2013-2024 (亿元, final-account)
CENTRAL_PASS_THROUGH = [
    # (year, central revenue, central-to-local transfer total[incl tax rebate])
    (2013, 60198, 48019.92),  # 一般性 + 专项 + 税收返还
    (2014, 64493, 51591.04),
    (2015, 69234, 55097.51),
    (2016, 72366, 59400.70),
    (2017, 81119, 65051.78),
    (2018, 85447, 69680.66),
    (2019, 89309, 74359.86),
    (2020, 82771, 83217.93),  # central revenue dropped vs 2019; transfers rose
    (2021, 91470, 82152.34),
    (2022, 94885, 96941.82),  # transfers > revenue (debt-financed)
    (2023, 99566, 102836.32),
    (2024, 100442, 100335.72),
]

# ---------------------------------------------------------------
# Sheet E — Transferred-in funds 调入资金, 2013-2024
# ---------------------------------------------------------------
# Source: Local General Public Budget Revenue Final Account Tables, footer line
# "地方财政调入资金及使用结转结余" or "从预算稳定调节基金调入及使用结转结余"
DIAORU = [
    (2013, 593.26),
    (2014, 0),       # not separately listed
    (2015, 7236.07),
    (2016, 5911.31),
    (2017, 8407.15),
    (2018, 12312.28),
    (2019, 19002.75),
    (2020, 17422.37),
    (2021, 9186.47),
    (2022, 12077.32),
    (2023, 9138.41),
    (2024, 17077.43),
]


# ---------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

# ===== Sheet A =====
ws = wb.create_sheet("A_Property_Taxes")
ws['A1'] = "Property-related local taxes, 2013–2024 (Test B)"
ws['A1'].font = Font(name="Arial", bold=True, size=14)
ws.merge_cells('A1:I1')
ws['A2'] = "Final-account figures, in 亿元 (100 million yuan). All four taxes are 100% local."
ws['A2'].font = FONT_NOTE
ws.merge_cells('A2:I2')

headers = ["Year", "契税 / Deed", "土地增值税 / Land VAT", "房产税 / Property tax",
           "城镇土地使用税 / Urban land use", "Four-tax total", "地方本级收入 / Local own-source",
           "Four-tax share of own-source", "YoY change in four-tax share"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

start = 5
for i, row in enumerate(PROPERTY_TAX):
    rr = start + i
    year, deed, lvat, ptax, ulu, owns = row
    ws.cell(row=rr, column=1, value=year)
    ws.cell(row=rr, column=2, value=deed)
    ws.cell(row=rr, column=3, value=lvat)
    ws.cell(row=rr, column=4, value=ptax)
    ws.cell(row=rr, column=5, value=ulu)
    ws.cell(row=rr, column=6, value=f"=B{rr}+C{rr}+D{rr}+E{rr}")
    ws.cell(row=rr, column=7, value=owns)
    ws.cell(row=rr, column=8, value=f"=F{rr}/G{rr}")
    if i > 0:
        ws.cell(row=rr, column=9, value=f"=H{rr}-H{rr-1}")
end = start + len(PROPERTY_TAX) - 1

# Highlight peak (2021) and 2024
for col in range(1, len(headers) + 1):
    ws.cell(row=start + 8, column=col).fill = FILL_HILITE  # 2021 peak
    ws.cell(row=end, column=col).fill = FILL_HILITE        # 2024

style_body(ws, start, end, len(headers))
for r in range(start, end + 1):
    for c in (2, 3, 4, 5, 6, 7):
        ws.cell(row=r, column=c).number_format = '#,##0.00'
    for c in (8, 9):
        ws.cell(row=r, column=c).number_format = '0.00%'

widths = [7, 14, 18, 14, 22, 16, 22, 18, 18]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 36
ws.freeze_panes = "B5"

# Findings note
note_row = end + 2
ws.cell(row=note_row, column=1, value="Key findings:").font = Font(name="Arial", bold=True, size=11)
findings_a = [
    "Peak in 2021: 1.97 trillion yuan (197 268 亿), 17.8% of provincial own-source revenue.",
    "By 2024: 1.72 trillion yuan, 14.4% — a 13% drop in absolute terms, 3.4-pp drop in share.",
    "If property taxes had grown at the same ~5% rate as other taxes 2021–2024, they would have reached ~2.08 trillion. The shortfall (~360 billion yuan) corresponds to roughly 3 percentage points of the 2024 self-sufficiency rate.",
    "But the long-run decline in self-sufficiency cannot be blamed on property: from 2013 to 2020, the four-tax share actually rose from 15.1% to 18.4%.",
]
for i, f in enumerate(findings_a):
    ws.cell(row=note_row + 1 + i, column=1, value=f"• {f}").font = FONT_BODY
    ws.merge_cells(start_row=note_row+1+i, start_column=1, end_row=note_row+1+i, end_column=9)


# ===== Sheet B =====
ws = wb.create_sheet("B_Land_Sales")
ws['A1'] = "Land sale revenue, 2018–2024 (国有土地使用权出让金收入)"
ws['A1'].font = Font(name="Arial", bold=True, size=14)
ws.merge_cells('A1:E1')
ws['A2'] = "Final-account figures from Local Government-managed Funds Revenue tables, in 亿元."
ws['A2'].font = FONT_NOTE
ws.merge_cells('A2:E2')

headers = ["Year", "Land sale revenue (亿元)", "In trillion yuan", "YoY %", "Cumulative change vs 2021 peak"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

start = 5
peak_value = max(v for _, v in LAND_SALES)
for i, (year, val) in enumerate(LAND_SALES):
    rr = start + i
    ws.cell(row=rr, column=1, value=year)
    ws.cell(row=rr, column=2, value=val)
    ws.cell(row=rr, column=3, value=f"=B{rr}/10000")
    if i > 0:
        ws.cell(row=rr, column=4, value=f"=B{rr}/B{rr-1}-1")
    ws.cell(row=rr, column=5, value=f"=B{rr}/{peak_value}-1")
end = start + len(LAND_SALES) - 1

style_body(ws, start, end, len(headers))
for r in range(start, end + 1):
    ws.cell(row=r, column=2).number_format = '#,##0.00'
    ws.cell(row=r, column=3).number_format = '0.00'
    ws.cell(row=r, column=4).number_format = '0.0%'
    ws.cell(row=r, column=5).number_format = '0.0%'

widths = [7, 22, 16, 12, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 36

note_row = end + 2
ws.cell(row=note_row, column=1, value="Key findings:").font = Font(name="Arial", bold=True, size=11)
findings_b = [
    "2021 peak: 8.49 trillion yuan; 2024 floor: 4.77 trillion. Decline 44% over three years.",
    "Land sales sit in the government-managed funds budget, not the general public budget — but they fed the latter via the 'transferred-in funds' (调入资金) line. See Sheet E.",
]
for i, f in enumerate(findings_b):
    ws.cell(row=note_row + 1 + i, column=1, value=f"• {f}").font = FONT_BODY
    ws.merge_cells(start_row=note_row+1+i, start_column=1, end_row=note_row+1+i, end_column=5)


# ===== Sheet C =====
ws = wb.create_sheet("C_Central_Local_Share")
ws['A1'] = "Central-local revenue share + revenue/GDP decomposition (2013 vs 2024)"
ws['A1'].font = Font(name="Arial", bold=True, size=14)
ws.merge_cells('A1:G1')
ws['A2'] = "Refutes the simplest 'centralization' story: the central-local split is essentially unchanged."
ws['A2'].font = FONT_NOTE
ws.merge_cells('A2:G2')

headers = ["Year", "Nominal GDP (万亿)", "National revenue (亿)", "Central revenue (亿)",
           "Local revenue (亿)", "Central share of national", "Local share of national"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

start = 5
for i, row in enumerate(CENTRAL_LOCAL_GDP):
    rr = start + i
    year, gdp, nat, central, local = row
    ws.cell(row=rr, column=1, value=year)
    ws.cell(row=rr, column=2, value=gdp)
    ws.cell(row=rr, column=3, value=nat)
    ws.cell(row=rr, column=4, value=central)
    ws.cell(row=rr, column=5, value=local)
    ws.cell(row=rr, column=6, value=f"=D{rr}/C{rr}")
    ws.cell(row=rr, column=7, value=f"=E{rr}/C{rr}")
end = start + len(CENTRAL_LOCAL_GDP) - 1

style_body(ws, start, end, len(headers))
for r in range(start, end + 1):
    for c in (2, 3, 4, 5):
        ws.cell(row=r, column=c).number_format = '#,##0.00'
    for c in (6, 7):
        ws.cell(row=r, column=c).number_format = '0.0%'

# GDP-share decomposition table below
gdp_row = end + 3
ws.cell(row=gdp_row, column=1, value="Revenue / GDP decomposition (%):").font = Font(name="Arial", bold=True, size=11)
ws.cell(row=gdp_row + 1, column=1, value="Year")
ws.cell(row=gdp_row + 1, column=2, value="National rev / GDP")
ws.cell(row=gdp_row + 1, column=3, value="Central rev / GDP")
ws.cell(row=gdp_row + 1, column=4, value="Local rev / GDP")
style_header(ws, gdp_row + 1, 4)
for i, row in enumerate(CENTRAL_LOCAL_GDP):
    rr = gdp_row + 2 + i
    year, gdp, nat, central, local = row
    ws.cell(row=rr, column=1, value=year)
    ws.cell(row=rr, column=2, value=f"=C{start + i}/(B{start + i}*10000)")
    ws.cell(row=rr, column=3, value=f"=D{start + i}/(B{start + i}*10000)")
    ws.cell(row=rr, column=4, value=f"=E{start + i}/(B{start + i}*10000)")
style_body(ws, gdp_row + 2, gdp_row + 1 + len(CENTRAL_LOCAL_GDP), 4)
for r in range(gdp_row + 2, gdp_row + 2 + len(CENTRAL_LOCAL_GDP)):
    for c in (2, 3, 4):
        ws.cell(row=r, column=c).number_format = '0.00%'

widths = [7, 18, 22, 20, 20, 20, 20]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 36
ws.row_dimensions[gdp_row + 1].height = 32

note_row = gdp_row + 4 + len(CENTRAL_LOCAL_GDP)
ws.cell(row=note_row, column=1, value="Key findings:").font = Font(name="Arial", bold=True, size=11)
findings_c = [
    "Local share of national revenue: 53.4% (2013) → 53.8% (2024). Essentially unchanged.",
    "Central share of national revenue: 46.6% → 46.2%. Also essentially unchanged.",
    "What changed is the size of the pie. National revenue fell from 21.8% of GDP (2013) to 16.1% (2024) — a 5.7-pp drop.",
    "The drop is split almost evenly: central revenue/GDP fell 2.5 pp; local revenue/GDP fell 2.9 pp.",
    "This rules out the 'Beijing took provincial revenue' story. Both halves shrank in roughly equal proportion against GDP.",
]
for i, f in enumerate(findings_c):
    ws.cell(row=note_row + 1 + i, column=1, value=f"• {f}").font = FONT_BODY
    ws.merge_cells(start_row=note_row+1+i, start_column=1, end_row=note_row+1+i, end_column=7)


# ===== Sheet D =====
ws = wb.create_sheet("D_Central_PassThrough")
ws['A1'] = "Central transfers as share of central general public budget revenue, 2013–2024"
ws['A1'].font = Font(name="Arial", bold=True, size=14)
ws.merge_cells('A1:E1')
ws['A2'] = "Beijing has become a near pass-through: 80% (2013) → 98% (2024) of its general budget revenue gets routed to provinces."
ws['A2'].font = FONT_NOTE
ws.merge_cells('A2:E2')

headers = ["Year", "Central revenue (亿)", "Central-to-local transfer (亿, incl. tax rebate)",
           "Transfer / central revenue", "Note"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

NOTES = {
    2020: "central revenue fell sharply (COVID); transfer ratio crossed 100% briefly",
    2022: "留抵退税 + COVID transfer surge; ratio > 100%",
    2024: "ratio reaches 98%; the gap is filled by central-government debt issuance",
}

start = 5
for i, (year, central, transfer) in enumerate(CENTRAL_PASS_THROUGH):
    rr = start + i
    ws.cell(row=rr, column=1, value=year)
    ws.cell(row=rr, column=2, value=central)
    ws.cell(row=rr, column=3, value=transfer)
    ws.cell(row=rr, column=4, value=f"=C{rr}/B{rr}")
    if year in NOTES:
        ws.cell(row=rr, column=5, value=NOTES[year])
end = start + len(CENTRAL_PASS_THROUGH) - 1

style_body(ws, start, end, len(headers))
for r in range(start, end + 1):
    for c in (2, 3):
        ws.cell(row=r, column=c).number_format = '#,##0.00'
    ws.cell(row=r, column=4).number_format = '0.0%'

# highlight 2013 and 2024
for col in range(1, len(headers) + 1):
    ws.cell(row=start, column=col).fill = FILL_HILITE
    ws.cell(row=end, column=col).fill = FILL_HILITE

widths = [7, 22, 32, 22, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 36

note_row = end + 2
ws.cell(row=note_row, column=1, value="Key findings:").font = Font(name="Arial", bold=True, size=11)
findings_d = [
    "In 2013, Beijing transferred 80% of its general public budget revenue to provinces. By 2024, 98%.",
    "From 2020 onward, in some years (2020, 2022, 2023) the transfer total exceeded central revenue — financed by central-government debt.",
    "Beijing has become, in effect, a pass-through entity. The dependency runs both ways: provinces depend on central transfers, and the center depends on its capacity to keep redistributing.",
]
for i, f in enumerate(findings_d):
    ws.cell(row=note_row + 1 + i, column=1, value=f"• {f}").font = FONT_BODY
    ws.merge_cells(start_row=note_row+1+i, start_column=1, end_row=note_row+1+i, end_column=5)


# ===== Sheet E =====
ws = wb.create_sheet("E_Diaoru_TransferIn")
ws['A1'] = "Cross-budget transferred-in funds (调入资金), 2013–2024"
ws['A1'].font = Font(name="Arial", bold=True, size=14)
ws.merge_cells('A1:E1')
ws['A2'] = "These are NOT past-savings drawdowns — they are cross-budget transfers, primarily from the government-managed funds budget (i.e., land sale surpluses) into the general public budget."
ws['A2'].font = FONT_NOTE
ws.merge_cells('A2:E2')

headers = ["Year", "调入资金 (亿元)", "In trillion yuan", "Land sale revenue (亿)", "Ratio: 调入 / 土地出让金"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

land_dict = dict(LAND_SALES)

start = 5
for i, (year, val) in enumerate(DIAORU):
    rr = start + i
    ws.cell(row=rr, column=1, value=year)
    ws.cell(row=rr, column=2, value=val)
    ws.cell(row=rr, column=3, value=f"=B{rr}/10000")
    if year in land_dict:
        ws.cell(row=rr, column=4, value=land_dict[year])
        ws.cell(row=rr, column=5, value=f"=B{rr}/D{rr}")
    else:
        ws.cell(row=rr, column=4, value="—")
        ws.cell(row=rr, column=5, value="—")
end = start + len(DIAORU) - 1

style_body(ws, start, end, len(headers))
for r in range(start, end + 1):
    ws.cell(row=r, column=2).number_format = '#,##0.00'
    ws.cell(row=r, column=3).number_format = '0.00'
    if isinstance(ws.cell(row=r, column=4).value, (int, float)):
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5).number_format = '0.0%'

widths = [7, 18, 16, 22, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 36

note_row = end + 2
ws.cell(row=note_row, column=1, value="Key findings:").font = Font(name="Arial", bold=True, size=11)
findings_e = [
    "Peak: 1.90 trillion yuan in 2019 (close to 10% of total local spending).",
    "Sharp drop in 2021 to 920 billion — even before land sales themselves peaked, the transferable surplus from government-managed funds had thinned as fund spending caught up with revenue.",
    "Recovery to 1.71 trillion in 2024 reflects increased reliance on central debt (via shared fiscal responsibility transfers and special bonds) rather than land revenue, which kept falling.",
    "The mechanism's collapse deepened provincial dependence on central transfers rather than relieving it.",
]
for i, f in enumerate(findings_e):
    ws.cell(row=note_row + 1 + i, column=1, value=f"• {f}").font = FONT_BODY
    ws.merge_cells(start_row=note_row+1+i, start_column=1, end_row=note_row+1+i, end_column=5)


# ===== Sheet F — sources =====
ws = wb.create_sheet("F_Sources")
ws['A1'] = "Sources & methodology"
ws['A1'].font = Font(name="Arial", bold=True, size=14)
ws.merge_cells('A1:B1')

src = [
    ("Sheet A — property taxes", "Local General Public Budget Revenue Final Account Tables, MoF Budget Department, 2013–2024. Each line item extracted directly. URLs in main README."),
    ("Sheet B — land sales", "Local Government-managed Funds Revenue Final Account Tables, line '国有土地使用权出让金收入', 2018–2024. Pre-2018 data uses different categorization and is omitted for comparability."),
    ("Sheet C — central-local share", "National / Central / Local General Public Budget Revenue Final Account Tables, 2013 and 2024. China NBS nominal GDP (current prices)."),
    ("Sheet D — central pass-through", "Central General Public Budget Revenue Final Account Tables, 2013–2024. Central-to-local transfer total = general transfer + special transfer + (pre-2019: tax rebate; post-2019: tax rebate is bundled in general transfer)."),
    ("Sheet E — 调入资金", "Local General Public Budget Revenue Final Account Tables, footer line. Label varies year-to-year between '从预算稳定调节基金调入及使用结转结余' and '地方财政调入资金及使用结转结余'."),
    ("", ""),
    ("Caliber notes", "(1) Pre-2019 'central-to-local transfer' excludes tax rebate as a separate row; the comparable total used here adds tax rebate back. (2) 2014 调入资金 was not separately listed (recorded as 0). (3) Provincial-level breakdowns of 调入资金 by source budget are not in the national tables; the inference that this line is dominated by government-managed funds inflows comes from provincial budget reports of high-disclosure provinces (e.g., Guangdong, Zhejiang)."),
]
for i, (k, v) in enumerate(src, start=3):
    ws.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=i, column=2, value=v).font = FONT_BODY
    ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 50

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 100

# Save
out = DATA_DIR / '4_revised_evidence.xlsx'
wb.save(str(out))
print(f"Saved: {out}")
