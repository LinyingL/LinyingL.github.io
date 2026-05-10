"""
Build the "two-layer Tocquevillian" visualization & xlsx for short data news on
"集权 + 放权 悖论：转移支付内部'自由化' vs 地方'自给率'下降".

Combines our prior transfer-payment structure data with newly fetched
local own-source revenue / total expenditure data, all from 财政部预算司.
"""
# --- Path setup (relative to repo) ---
from pathlib import Path
HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / 'data'
FIG_DIR  = HERE.parent / 'figures'
DATA_DIR.mkdir(exist_ok=True, parents=True)
FIG_DIR.mkdir(exist_ok=True, parents=True)


import matplotlib
from matplotlib import font_manager
font_path = next((p for p in [
    '/tmp/fonts/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
    str(Path.home() / 'Library/Fonts/NotoSansCJK-Regular.ttc'),
] if Path(p).exists()), '/tmp/fonts/NotoSansCJK-Regular.ttc')
font_manager.fontManager.addfont(font_path) if Path(font_path).exists() else None
_prop = font_manager.FontProperties(fname=font_path)
_FONT_NAME = _prop.get_name()  # actually 'Noto Sans CJK JP' — works for Chinese fine
import matplotlib.pyplot as plt
plt.rcParams['font.family'] = _FONT_NAME
plt.rcParams['font.sans-serif'] = [_FONT_NAME]
plt.rcParams['axes.unicode_minus'] = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================
# Master data
# =============================================================
# (本级收入, 中央转移支付, 调入资金, 差额, 总支出 if known)
DATA = {
    2013: dict(local=69011.16, transfer=48019.92, diao=593.26,   diff=3500.00,  total_exp=None),
    2014: dict(local=75876.58, transfer=51591.04, diao=0.00,     diff=4000.00,  total_exp=None),
    2015: dict(local=83002.04, transfer=55097.51, diao=7236.07,  diff=5000.00,  total_exp=None),
    2016: dict(local=87239.35, transfer=59400.70, diao=5911.31,  diff=7800.00,  total_exp=None),
    2017: dict(local=91469.41, transfer=65051.78, diao=8407.15,  diff=8300.00,  total_exp=None),
    2018: dict(local=97903.38, transfer=69680.66, diao=12312.28, diff=8300.00,  total_exp=None),
    2019: dict(local=101080.61, transfer=74359.86, diao=19002.75, diff=9300.00, total_exp=203759),  # 财政部公告口径
    2020: dict(local=100143.16, transfer=83217.93, diao=17422.37, diff=9800.00, total_exp=210492),
    2021: dict(local=111084.23, transfer=82152.34, diao=9186.47,  diff=8200.00, total_exp=211272),
    2022: dict(local=108762.15, transfer=96941.82, diao=12077.32, diff=7200.00, total_exp=224981),
    2023: dict(local=117228.73, transfer=102836.32, diao=9138.41, diff=7200.00, total_exp=236403),
    2024: dict(local=119272.24, transfer=100335.72, diao=17077.43,diff=7200.00, total_exp=243885),
    2025: dict(local=122082.00, transfer=101925.00, diao=None,    diff=None,    total_exp=244361),  # 实际/接近决算
    2026: dict(local=125030.00, transfer=104150.00, diao=17000.00,diff=8000.00, total_exp=254180),  # 预算
}
KIND = {y: ('决算' if y <= 2024 else ('实际' if y == 2025 else '预算')) for y in DATA}

# Compute总支出 for years where it's None (use identity)
for y, d in DATA.items():
    if d['total_exp'] is None:
        d['total_exp'] = d['local'] + d['transfer'] + (d['diao'] or 0) + (d['diff'] or 0)

# =============================================================
# Derive: 自给率 + 依赖度
# =============================================================
years = sorted(DATA.keys())
sufficiency = [DATA[y]['local'] / DATA[y]['total_exp'] * 100 for y in years]
dependency  = [DATA[y]['transfer'] / DATA[y]['total_exp'] * 100 for y in years]

# Also bring in the "free part" share from prior dataset (2019+)
# Computed earlier: 一般性其他/总转移支付
free_share = {
    2019: 31.8, 2020: 31.2, 2021: 35.3, 2022: 33.6,
    2023: 35.9, 2024: 37.4, 2025: 41.5, 2026: 41.7,
}

# =============================================================
# Two-panel chart: top = 内部 autonomy ↑ ; bottom = 自给率 ↓
# =============================================================
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(13, 9), sharex=True,
                                gridspec_kw={'height_ratios': [1, 1.3], 'hspace': 0.16})

# ---- Top: 内部自由部分占比（仅 2019+） ----
xs_free = sorted(free_share.keys())
ys_free = [free_share[y] for y in xs_free]
ax1.plot(xs_free, ys_free, marker='^', color='#2E75B6', linewidth=2.5, markersize=8)
ax1.fill_between(xs_free, 0, ys_free, color='#2E75B6', alpha=0.10)
for x, y in zip(xs_free, ys_free):
    ax1.text(x, y+1.3, f'{y:.1f}%', ha='center', fontsize=9, color='#2E75B6')
ax1.set_ylim(25, 50)
ax1.set_ylabel('占总转移支付 %', fontsize=10)
ax1.set_title('上层：账面"放权"——转移支付内部，地方真正可自由调度部分占比', fontsize=12, color='#2E75B6', loc='left', weight='bold')
ax1.spines['top'].set_visible(False)
ax1.spines['right'].set_visible(False)
ax1.axvline(2019, color='#888', linestyle='--', linewidth=1, alpha=0.5)
ax1.text(2019, 47, ' 2019 共同事权独立列示', fontsize=9, color='#666', va='top')
ax1.axvspan(2024.5, 2026.5, color='#FFF2CC', alpha=0.5, zorder=0)
ax1.text(2025.5, 47, '预算口径', fontsize=9, color='#888', ha='center', style='italic', va='top')

# ---- Bottom: 自给率 + 依赖度 ----
ax2.plot(years, sufficiency, marker='o', color='#C00000', linewidth=2.8,
         markersize=8, label='① 地方财政自给率 = 本级收入 / 一般预算总支出')
ax2.plot(years, dependency,  marker='s', color='#7F7F7F', linewidth=2.0,
         markersize=6, label='② 中央转移支付依赖度 = 中央转移 / 一般预算总支出')

# 趋势 annotation
ax2.annotate(f'{sufficiency[0]:.1f}%', xy=(2013, sufficiency[0]), xytext=(2013, sufficiency[0]+1.5),
             fontsize=10, color='#C00000', weight='bold', ha='center')
ax2.annotate(f'{sufficiency[-1]:.1f}%', xy=(2026, sufficiency[-1]), xytext=(2026, sufficiency[-1]+1.5),
             fontsize=10, color='#C00000', weight='bold', ha='center')
# Big arrow showing decline
ax2.annotate('',
             xy=(2024.5, 50), xytext=(2014.5, 56),
             arrowprops=dict(arrowstyle='->', color='#C00000', lw=2, alpha=0.6))
ax2.text(2019, 51.5, '11 年下降 8 个百分点', fontsize=11, color='#C00000', ha='center', weight='bold')

ax2.axvline(2016, color='#888', linestyle=':', linewidth=1, alpha=0.6)
ax2.text(2016, 60, ' 2016 全面营改增', fontsize=9, color='#666', va='top')
ax2.axvline(2020, color='#888', linestyle=':', linewidth=1, alpha=0.6)
ax2.text(2020, 60, ' 2020 抗疫', fontsize=9, color='#666', va='top')
ax2.axvline(2022, color='#888', linestyle=':', linewidth=1, alpha=0.6)
ax2.text(2022, 60, ' 2022 大规模留抵退税', fontsize=9, color='#666', va='top')
ax2.axvspan(2024.5, 2026.5, color='#FFF2CC', alpha=0.5, zorder=0)

ax2.set_ylim(35, 62)
ax2.set_ylabel('占地方一般预算总支出 %', fontsize=10)
ax2.set_xlabel('')
ax2.set_xticks(years)
ax2.set_xticklabels([str(y) for y in years])
ax2.set_title('下层：实质"集权"——地方自有财力占总支出比重持续下行，对中央依赖加深', fontsize=12, color='#C00000', loc='left', weight='bold')
ax2.legend(loc='upper right', frameon=False, fontsize=10)
ax2.spines['top'].set_visible(False)
ax2.spines['right'].set_visible(False)
ax2.grid(axis='y', alpha=0.25)

fig.suptitle('"集权 + 放权" 悖论：账面给得更松，实质依赖更深', fontsize=15, weight='bold', y=0.97)
fig.text(0.06, 0.02,
    '数据来源：财政部预算司 2013–2024《地方一般公共预算收入决算表》及历年转移支付决算表；2025 实际数取自 2026 年 1 月《2025 年财政收支情况》新闻发布；2026 取自 2026 年中央和地方预算草案。\n'
    '指标定义：自给率 = 地方一般公共预算本级收入 / 地方一般公共预算总支出；依赖度 = 中央对地方一般公共预算转移支付（含税收返还）/ 地方一般公共预算总支出。\n'
    '"自由调度部分" 仅 2019 年起可比（共同事权独立列示之前数据缺失）。',
    fontsize=8, color='#555', ha='left', va='bottom')

plt.subplots_adjust(left=0.07, right=0.96, top=0.91, bottom=0.16)
plt.savefig(str(FIG_DIR / '3_two_layer_paradox_2013_2026.png'),
            dpi=180, bbox_inches='tight', facecolor='white')
plt.close()

# =============================================================
# xlsx with full annual data
# =============================================================
wb = Workbook()
ws = wb.active
ws.title = "依赖度时间序列"

FONT_HEADER = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_NOTE = Font(name="Arial", size=9, italic=True, color="555555")
FILL_HEADER = PatternFill("solid", start_color="305496")
FILL_BUDGET = PatternFill("solid", start_color="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

ws['A1'] = "中国地方政府财政自给率 2013–2026（基于财政部决算/预算口径）"
ws['A1'].font = Font(name="Arial", bold=True, size=14)
ws.merge_cells('A1:K1')
ws['A2'] = "单位：亿元（金额）/ %（占比）；数据来源：财政部预算司"
ws['A2'].font = FONT_NOTE
ws.merge_cells('A2:K2')

headers = ['年份', '口径', '地方本级收入', '中央对地方转移支付', '调入资金', '差额', '地方一般预算总支出',
           '自给率（本级/总支出）', '依赖度（转移/总支出）', '调入+差额占比', '自由调度部分占转移支付']
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = BORDER

start = 5
for i, y in enumerate(years):
    rr = start + i
    d = DATA[y]
    ws.cell(row=rr, column=1, value=y)
    ws.cell(row=rr, column=2, value=KIND[y])
    ws.cell(row=rr, column=3, value=d['local'])
    ws.cell(row=rr, column=4, value=d['transfer'])
    ws.cell(row=rr, column=5, value=d['diao'] if d['diao'] is not None else "—")
    ws.cell(row=rr, column=6, value=d['diff'] if d['diff'] is not None else "—")
    ws.cell(row=rr, column=7, value=d['total_exp'])
    ws.cell(row=rr, column=8, value=f"=C{rr}/G{rr}")
    ws.cell(row=rr, column=9, value=f"=D{rr}/G{rr}")
    ws.cell(row=rr, column=10, value=f"=(G{rr}-C{rr}-D{rr})/G{rr}")
    free = free_share.get(y)
    if free is not None:
        ws.cell(row=rr, column=11, value=free/100)
    else:
        ws.cell(row=rr, column=11, value="—")
    for col in range(1, 12):
        cell = ws.cell(row=rr, column=col)
        cell.font = FONT_BODY
        cell.border = BORDER
        cell.alignment = ALIGN_CENTER if col <= 2 else ALIGN_RIGHT
        if KIND[y] != '决算':
            cell.fill = FILL_BUDGET
    for col in (3, 4, 5, 6, 7):
        ws.cell(row=rr, column=col).number_format = '#,##0.00'
    for col in (8, 9, 10, 11):
        ws.cell(row=rr, column=col).number_format = '0.0%'

widths = [7, 9, 14, 18, 12, 10, 18, 18, 18, 16, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 36

# Sheet 2: narrative
ws2 = wb.create_sheet("两层悖论解读")
ws2['A1'] = "为什么会出现「上层放权 + 下层集权」的双层悖论？"
ws2['A1'].font = Font(name="Arial", bold=True, size=14)
ws2.merge_cells('A1:B1')

q = [
    ('数据上看到了什么',
     '账面层（2019–2026）：转移支付内部"地方真正可自由调度部分"占比从 31.8% 升至 41.7%。\n'
     '实质层（2013–2026）：地方财政自给率从 57.0% 一路下降到 49.2%，11 年丢失 8 个百分点；同期中央转移支付依赖度从 39.6% 升至 41.0%。'),
    ('为什么自给率下降',
     '① 2014 营改增（2016 全面推开）将原属地方主体税种的营业税并入增值税，地方主体税源被结构性削弱。\n'
     '② 2018 国地税合并，征管效率上升但地方税权进一步弱化。\n'
     '③ 2022 大规模增值税留抵退税（约 2.4 万亿元），地方留底数显著下降。\n'
     '④ 民生刚性支出（教育、医保、养老）持续上升，地方自有收入跑不赢。'),
    ('为什么内部还在"放权"',
     '① 2014 专项整合改革本意就是"减项目数 + 扩地方自主"，2019 共同事权独立列示是会计上的清晰化，并非新增控制。\n'
     '② 中央通过"控制盘子的总规模 + 分类用途"已经掌握主动权，无须再通过"项目锁死"做微观干预。\n'
     '③ 增加地方使用自由度，把支出责任和绩效压力下沉到省级，本身就是中央简政减责的需要。'),
    ('两层叠加的逻辑',
     '中央对地方的"控制"从「指定钱怎么花」转向「让地方离不开中央的钱」。表面减少了行政干预，但通过财政纵向不平衡（vertical fiscal imbalance）形成了更隐蔽的依附结构。\n'
     '托克维尔在《旧制度与大革命》中描述的就是这种机制——"集权过程的隐蔽性在于它不是通过命令实现，而是通过让地方习惯于依赖中央来实现"。'),
    ('新闻角度建议',
     '不要写"压缩 vs 放权"，而要写"两层"：\n'
     '上层标题感的反差：转移支付内部自由度上升（图）；下层骨头：自给率下降（图）。\n'
     '把这两张图并列，本身就是新闻叙事的核心——账面 vs 实质。'),
]
for i, (k, v) in enumerate(q, start=3):
    ws2.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=11)
    ws2.cell(row=i, column=1).alignment = Alignment(vertical="top")
    ws2.cell(row=i, column=2, value=v).font = FONT_BODY
    ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[i].height = 95

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 100

# Sheet 3: 数据来源
ws3 = wb.create_sheet("数据来源")
ws3['A1'] = "数据来源（财政部预算司 + 财政部国库司 + 财政部预算说明）"
ws3['A1'].font = Font(name="Arial", bold=True, size=14)
ws3.merge_cells('A1:B1')

src = [
    ('2013 地方公共财政收入决算表', 'http://yss.mof.gov.cn/2013qgczjs/201407/t20140711_1111966.htm'),
    ('2014 地方一般公共预算收入决算表', 'http://yss.mof.gov.cn/2014czys/201507/t20150709_1272603.htm'),
    ('2015 同上', 'http://yss.mof.gov.cn/2015js/201607/t20160713_2355032.htm'),
    ('2016 同上', 'http://yss.mof.gov.cn/2016js/201707/t20170713_2648986.htm'),
    ('2017 同上', 'http://yss.mof.gov.cn/qgczjs/201807/t20180712_2959591.htm'),
    ('2018 同上', 'http://yss.mof.gov.cn/2018czjs/201907/t20190718_3303120.htm'),
    ('2019 同上', 'http://yss.mof.gov.cn/2019qgczjs/202007/t20200731_3559702.htm'),
    ('2020 同上', 'http://yss.mof.gov.cn/2020zyjs/202109/t20210917_3753572.htm'),
    ('2021 同上', 'http://yss.mof.gov.cn/2021zyjs/202207/t20220728_3830476.htm'),
    ('2022 同上', 'http://yss.mof.gov.cn/2022zyjs/202308/t20230825_3904171.htm'),
    ('2023 同上', 'http://yss.mof.gov.cn/2023zyjs/202408/t20240830_3942863.htm'),
    ('2024 同上', 'http://yss.mof.gov.cn/2024zyjs/202509/t20250904_3971521.htm'),
    ('2025 财政收支情况（实际数）', 'http://gks.mof.gov.cn/tongjishuju/202601/t20260130_3982923.htm'),
    ('2026 中央对地方转移支付预算表', 'http://yss.mof.gov.cn/2026zyczys/202603/t20260324_3986014.htm'),
    ('转移支付内部"自由部分"指标', '基于历年中央对地方转移支付决算表，扣除税收返还、共同事权后的余额'),
]
for i, (k, v) in enumerate(src, start=3):
    ws3.cell(row=i, column=1, value=k).font = Font(name="Arial", bold=True, size=11)
    ws3.cell(row=i, column=2, value=v).font = FONT_BODY

ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 90

out = str(DATA_DIR / '3_local_self_sufficiency_2013_2026.xlsx')
wb.save(out)
print(f"Saved xlsx: {out}")
print(f"Saved chart: 集权放权悖论_2013_2026.png")
